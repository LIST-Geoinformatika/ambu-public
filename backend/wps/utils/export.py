from io import BytesIO

from openpyxl import Workbook


def export_permit_to_xlsx(obj):
    # Create a new workbook
    workbook = Workbook()

    # Get the active worksheet
    sheet = workbook.active

    sheet.cell(row=1, column=1, value="PERMIT - {}".format(str(obj.uid)))

    sheet.cell(row=2, column=1, value="Submitted By:")
    sheet.cell(row=2, column=2, value="{} ({})".format(obj.submitted_by.full_name, obj.submitted_by.email))

    sheet.cell(row=3, column=1, value="Submitted On:")
    sheet.cell(row=3, column=2, value=obj.submitted_on.strftime('%d.%m.%Y %H:%M'))

    sheet.cell(row=4, column=1, value="Operator name:")
    sheet.cell(row=4, column=2, value=obj.operator_name)

    sheet.cell(row=5, column=1, value="NACE:")
    if obj.nace_code:
        sheet.cell(row=5, column=2, value="{} - {}".format(obj.nace_code.code, obj.nace_code.description))

    sheet.cell(row=7, column=1, value="Validated by:")
    if obj.validated_by:
        sheet.cell(row=7, column=2, value=obj.validated_by.full_name)

    sheet.cell(row=8, column=1, value="Validated on:")
    if obj.validated_on:
        sheet.cell(row=8, column=2, value=obj.validated_on.strftime('%d.%m.%Y %H:%M'))

    sheet.cell(row=9, column=1, value="Status:")
    sheet.cell(row=9, column=2, value=obj.status)

    sheet.cell(row=10, column=1, value="Remark:")
    sheet.cell(row=10, column=2, value=obj.remark)

    sheet.cell(row=12, column=1, value="ABSTRACTION POINT")
    sheet.cell(row=13, column=1, value="Identifier")
    sheet.cell(row=13, column=2, value="Location")
    sheet.cell(row=13, column=3, value="Sub-basin")
    sheet.cell(row=13, column=4, value="Watercourse name")
    sheet.cell(row=13, column=5, value="WB code")
    sheet.cell(row=13, column=6, value="Approved")

    ap = obj.abstraction_points.last()
    if ap:
        sheet.cell(row=14, column=1, value=ap.identifier)
        sheet.cell(row=14, column=2, value='{}, {}'.format(ap.geom.coords[0], ap.geom.coords[1]))

        if ap.subbasin:
            sheet.cell(row=14, column=3, value=ap.subbasin.name)

        if ap.water_body:
            sheet.cell(row=14, column=4, value=ap.water_body.name)
            sheet.cell(row=14, column=5, value=ap.water_body.wb_code)

        sheet.cell(row=14, column=6, value=ap.approved)

    sheet.cell(row=16, column=1, value="DISCHARGE POINT")
    sheet.cell(row=17, column=1, value="Identifier")
    sheet.cell(row=17, column=2, value="Location")
    sheet.cell(row=17, column=3, value="Sub-basin")
    sheet.cell(row=17, column=4, value="Watercourse name")
    sheet.cell(row=17, column=5, value="WB code")
    sheet.cell(row=17, column=6, value="Approved")

    dp = obj.discharge_points.last()
    if dp:
        sheet.cell(row=17, column=1, value=dp.identifier)
        sheet.cell(row=17, column=2, value='{}, {}'.format(dp.geom.coords[0], dp.geom.coords[1]))

        if dp.subbasin:
            sheet.cell(row=17, column=3, value=dp.subbasin.name)
        if dp.water_body:
            sheet.cell(row=17, column=4, value=dp.water_body.name)
            sheet.cell(row=17, column=5, value=dp.water_body.wb_code)

        sheet.cell(row=17, column=6, value=dp.approved)

    sheet.cell(row=20, column=1, value="Time per month")
    sheet.cell(row=21, column=1, value="Abstraction m3")
    sheet.cell(row=22, column=1, value="Abstraction m3/s")
    sheet.cell(row=23, column=1, value="Discharge m3")
    sheet.cell(row=24, column=1, value="Discharge m3/s")

    months = [
        {1: 'January'},
        {2: 'February'},
        {3: 'March'},
        {4: 'April'},
        {5: 'May'},
        {6: 'June'},
        {7: 'July'},
        {8: 'August'},
        {9: 'September'},
        {10: 'October'},
        {11: 'November'},
        {12: 'December'}
    ]

    for month_dict in months:
        for month_num, month_name in month_dict.items():
            sheet.cell(row=19, column=1+month_num, value=month_name)

            if obj.time_per_month:
                sheet.cell(row=20, column=1+month_num, value=obj.time_per_month[str(month_num)])

            if obj.abstraction_m3_per_month:
                sheet.cell(row=21, column=1+month_num, value=obj.abstraction_m3_per_month[str(month_num)])

            if obj.abstraction_m3s_per_month:
                sheet.cell(row=22, column=1+month_num, value=obj.abstraction_m3s_per_month[str(month_num)])

            if obj.discharge_m3_per_month:
                sheet.cell(row=23, column=1+month_num, value=obj.discharge_m3_per_month[str(month_num)])

            if obj.discharge_m3s_per_month:
                sheet.cell(row=24, column=1+month_num, value=obj.discharge_m3s_per_month[str(month_num)])

    # sheet.cell(row=25, column=1, value="EV 1")
    # sheet.cell(row=26, column=1, value="EV 2")
    # sheet.cell(row=27, column=1, value="EV 3")
    # sheet.cell(row=28, column=1, value="EV 1 2 3")

    # Save the workbook to a byte stream
    output = BytesIO()
    workbook.save(output)

    return output
